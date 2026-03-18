#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1_letter_register.py — Регистрация входящей и исходящей корреспонденции
========================================================================
Базируется на версии monolith от 16.03.2026 (e09dd7a), часть «Регистрация писем».
Добавлена поддержка Word (.doc/.docx) и Excel (.xls/.xlsx) помимо PDF.

Входящие поля LanDocs (Tab от начала):
  0  — № вх      3  — файл     4  — Корреспондент
  5  — Дата       6  — № письма  8  — Подписант
  10 — Тема       15 — Связанное

Исходящие поля LanDocs (Tab от начала):
  0  — № письма   1  — Дата     5  — Тема
  6  — Исполнитель 9 — Компании  10 — ФИО получателей  16 — Связанное
"""

import os
import re
import sys
import json
import time
import shutil
import getpass
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

# ── Зависимости ──────────────────────────────────────────────────────────────

try:
    import win32api, win32con, win32clipboard
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

try:
    import openpyxl
    from openpyxl.styles import Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Конфигурация ──────────────────────────────────────────────────────────────

EXCEL_PATH_IN  = r"\\Prim-fs-serv\primrdu\СРЗА\Журнал регистрации корреспонденции\Журнал регистрации входящей документации.xlsx"
EXCEL_PATH_OUT = r"\\Prim-fs-serv\primrdu\СРЗА\Журнал регистрации корреспонденции\Журнал регистрации исходящей документации.xlsx"
DEFAULT_SAVE_FOLDER = r"\\Prim-fs-serv\primrdu\СРЗА\Дела СРЗА\19 Переписка"

TAB_DELAY  = 0.07
COPY_DELAY = 0.12

# ── Форматы писем ─────────────────────────────────────────────────────────────

LETTER_EXTS = ('.pdf', '.doc', '.docx', '.xls', '.xlsx')
LETTER_FILETYPES = [
    ("Письма (PDF, Word, Excel)", "*.pdf *.doc *.docx *.xls *.xlsx"),
    ("PDF файлы",                 "*.pdf"),
    ("Word файлы",                "*.doc *.docx"),
    ("Excel файлы",               "*.xls *.xlsx"),
    ("Все файлы",                 "*.*"),
]

# ── Настройки ─────────────────────────────────────────────────────────────────

def _appdata_dir() -> str:
    """Возвращает %APPDATA%\\DocsRegister, создаёт если нет."""
    appdata = os.environ.get('APPDATA') or os.path.join(
        os.environ.get('USERPROFILE', ''), 'AppData', 'Roaming')
    d = os.path.join(appdata, 'DocsRegister')
    os.makedirs(d, exist_ok=True)
    return d

SETTINGS_FILE = os.path.join(_appdata_dir(), 'letter_settings.json')

_settings: dict = {
    'org_abbreviations': {},
    'registrar_name': '',
}

def load_settings():
    global _settings
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                _settings.update(data)
    except Exception:
        pass

def save_settings():
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(_settings, f, ensure_ascii=False, indent=2)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить настройки:\n{e}")

# ── Клавиатура / буфер обмена ────────────────────────────────────────────────

def _clear_clipboard():
    if not HAS_WIN32: return
    try:
        win32clipboard.OpenClipboard(); win32clipboard.EmptyClipboard()
    except Exception: pass
    finally:
        try: win32clipboard.CloseClipboard()
        except Exception: pass

def _get_clipboard() -> str:
    if not HAS_WIN32: return ""
    try:
        win32clipboard.OpenClipboard()
        try:
            if win32clipboard.IsClipboardFormatAvailable(win32con.CF_UNICODETEXT):
                return (win32clipboard.GetClipboardData(win32con.CF_UNICODETEXT) or "").strip()
        finally:
            win32clipboard.CloseClipboard()
    except Exception: pass
    return ""

def _send_tab():
    if not HAS_WIN32: return
    win32api.keybd_event(win32con.VK_TAB, 0, 0, 0)
    win32api.keybd_event(win32con.VK_TAB, 0, win32con.KEYEVENTF_KEYUP, 0)

def _send_shift_tab():
    if not HAS_WIN32: return
    win32api.keybd_event(win32con.VK_SHIFT, 0, 0, 0)
    win32api.keybd_event(win32con.VK_TAB, 0, 0, 0)
    win32api.keybd_event(win32con.VK_TAB, 0, win32con.KEYEVENTF_KEYUP, 0)
    win32api.keybd_event(win32con.VK_SHIFT, 0, win32con.KEYEVENTF_KEYUP, 0)

def _send_ctrl_a():
    if not HAS_WIN32: return
    win32api.keybd_event(win32con.VK_CONTROL, 0, 0, 0)
    win32api.keybd_event(ord('A'), 0, 0, 0)
    win32api.keybd_event(ord('A'), 0, win32con.KEYEVENTF_KEYUP, 0)
    win32api.keybd_event(win32con.VK_CONTROL, 0, win32con.KEYEVENTF_KEYUP, 0)

def _send_ctrl_c():
    if not HAS_WIN32: return
    win32api.keybd_event(win32con.VK_CONTROL, 0, 0, 0)
    win32api.keybd_event(ord('C'), 0, 0, 0)
    win32api.keybd_event(ord('C'), 0, win32con.KEYEVENTF_KEYUP, 0)
    win32api.keybd_event(win32con.VK_CONTROL, 0, win32con.KEYEVENTF_KEYUP, 0)

def read_current_field() -> str:
    _clear_clipboard()
    time.sleep(0.04); _send_ctrl_a(); time.sleep(0.04)
    _send_ctrl_c(); time.sleep(COPY_DELAY)
    return _get_clipboard()

def navigate_tabs(n: int):
    for _ in range(n):
        _send_tab(); time.sleep(TAB_DELAY)

# ── Парсинг LanDocs ──────────────────────────────────────────────────────────

def extract_landocs_data_in() -> dict:
    data = {}
    current = 0
    _send_tab(); time.sleep(TAB_DELAY)
    _send_shift_tab(); time.sleep(TAB_DELAY)
    data['incoming_num'] = read_current_field()
    navigate_tabs(3 - current); current = 3; data['file_link'] = read_current_field()
    navigate_tabs(4 - current); current = 4; data['correspondent'] = read_current_field()
    navigate_tabs(5 - current); current = 5; data['date'] = read_current_field()
    navigate_tabs(6 - current); current = 6; data['letter_num'] = read_current_field()
    navigate_tabs(8 - current); current = 8; data['signatory'] = read_current_field()
    navigate_tabs(10 - current); current = 10; data['subject'] = read_current_field()
    navigate_tabs(15 - current); data['related'] = read_current_field()
    return data

def extract_landocs_data_out() -> dict:
    data = {}
    current = 0
    _send_tab(); time.sleep(TAB_DELAY)
    _send_shift_tab(); time.sleep(TAB_DELAY)
    data['letter_num'] = read_current_field()
    navigate_tabs(1 - current); current = 1; data['date'] = read_current_field()
    navigate_tabs(5 - current); current = 5; data['subject'] = read_current_field()
    navigate_tabs(6 - current); current = 6; data['executor'] = read_current_field()
    # Позиция 9 — компании, позиция 10 — ФИО (в LanDocs именно такой порядок)
    navigate_tabs(9 - current); current = 9; data['recipient_companies'] = read_current_field()
    navigate_tabs(10 - current); current = 10; data['recipient_names'] = read_current_field()
    navigate_tabs(16 - current); data['related'] = read_current_field()
    return data

# ── Утилиты ──────────────────────────────────────────────────────────────────

def find_latest_in_viewdir() -> str:
    local_app = os.environ.get('LOCALAPPDATA') or os.path.join(
        os.environ.get('USERPROFILE', ''), 'AppData', 'Local')
    view_dir = os.path.join(local_app, 'Temp', 'ViewDir')
    if not os.path.isdir(view_dir): return ''
    latest_path, latest_mtime = '', 0.0
    for dirpath, _, filenames in os.walk(view_dir):
        for fname in filenames:
            fpath = os.path.join(dirpath, fname)
            try:
                mtime = os.path.getmtime(fpath)
                if mtime > latest_mtime:
                    latest_mtime, latest_path = mtime, fpath
            except OSError: pass
    return latest_path

def sanitize_for_filename(text: str) -> str:
    return re.sub(r'[<>:"/\\|?*\r\n\t]', '_', text)

def parse_date(date_str: str):
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y.%m.%d'):
        try: return datetime.strptime(date_str.strip(), fmt)
        except ValueError: continue
    return None

def fmt_date_ymd(date_str: str) -> str:
    dt = parse_date(date_str); return dt.strftime('%Y-%m-%d') if dt else date_str

def fmt_date_dmy(date_str: str) -> str:
    dt = parse_date(date_str); return dt.strftime('%d.%m.%Y') if dt else date_str

def fmt_date_dmy_underscore(date_str: str) -> str:
    dt = parse_date(date_str)
    return dt.strftime('%d_%m_%Y') if dt else re.sub(r'[.\-/]', '_', date_str)

def abbreviate_fio(fio: str) -> str:
    """Фамилия Имя Отчество -> Фамилия И.О. (если >= 3 слова)"""
    parts = fio.strip().split()
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1][0]}.{parts[2][0]}."
    elif len(parts) == 2:
        return f"{parts[0]} {parts[1][0]}."
    return fio

def abbreviate_org(org: str) -> str:
    """Возвращает сокращённое название если есть в словаре, иначе исходное."""
    org_stripped = org.strip()
    abbrs = _settings.get('org_abbreviations', {})
    return abbrs.get(org_stripped, org_stripped)

def build_default_filename_in(date_str: str, incoming_num: str, letter_num: str,
                               subject: str = '', correspondent: str = '') -> str:
    inner = f"{incoming_num}_{sanitize_for_filename(letter_num)}_{fmt_date_dmy_underscore(date_str)}"
    subject_clean = sanitize_for_filename(subject.strip()) if subject.strip() else ''
    # Вставляем сокращённое название перед темой (если оно отличается от полного)
    abbr = ''
    if correspondent:
        a = abbreviate_org(correspondent)
        if a != correspondent.strip():
            abbr = sanitize_for_filename(a)
    after_from_parts = [p for p in [abbr, subject_clean] if p]
    after_from = ' '.join(after_from_parts)
    if after_from:
        return f"{fmt_date_ymd(date_str)} от {after_from} ({inner})"
    return f"{fmt_date_ymd(date_str)} ({inner})"

def build_default_filename_out(date_str: str, letter_num: str,
                               companies_str: str = '', subject: str = '') -> str:
    inner = f"{sanitize_for_filename(letter_num)}_{fmt_date_dmy_underscore(date_str)}"
    companies = [abbreviate_org(c.strip()) for c in companies_str.split(';') if c.strip()]
    company_part = ', '.join(sanitize_for_filename(c) for c in companies if c)
    subject_clean = sanitize_for_filename(subject.strip()) if subject.strip() else ''
    middle_parts = [p for p in [company_part, subject_clean] if p]
    if middle_parts:
        return f"{fmt_date_ymd(date_str)} в {' '.join(middle_parts)} ({inner})"
    return f"{fmt_date_ymd(date_str)} ({inner})"

def build_recipient_string(names_str: str, companies_str: str) -> str:
    names        = [abbreviate_fio(n.strip()) for n in names_str.split(';') if n.strip()]
    companies_raw = [c.strip() for c in companies_str.split(';') if c.strip()]
    companies_abbr = [abbreviate_org(c) for c in companies_raw]
    parts = []
    for i, (n, c_raw, c_abbr) in enumerate(zip(names, companies_raw, companies_abbr), 1):
        if c_abbr != c_raw:   # есть сокращённое название
            parts.append(f"{c_abbr} - {n}")
        else:
            parts.append(f"{i}. {n} {c_abbr}")
    return ';\n'.join(parts) if parts else names_str

def calc_folder_num(full_path: str) -> str:
    base = DEFAULT_SAVE_FOLDER.rstrip('\\/')
    norm = full_path.replace('/', '\\')
    if norm.lower().startswith(base.lower()):
        return norm[len(base):].lstrip('\\/')
    return norm

# ── Поиск в журнале входящих ─────────────────────────────────────────────────

def lookup_incoming_journal(incoming_num: str, year: str) -> dict | None:
    if not HAS_OPENPYXL: raise RuntimeError("openpyxl не установлен")
    if not os.path.exists(EXCEL_PATH_IN):
        raise FileNotFoundError(f"Журнал не найден:\n{EXCEL_PATH_IN}")
    wb = openpyxl.load_workbook(EXCEL_PATH_IN, read_only=True, data_only=True)
    service = {'служебный', 'шаблон', 'template'}
    year_sheets = [s for s in wb.sheetnames if s.lower() not in service]
    ws = None
    for s in wb.sheetnames:
        if s.strip() == year.strip(): ws = wb[s]; break
    if ws is None:
        ws = wb[year_sheets[-1]] if year_sheets else wb.worksheets[-1]
    num_clean = incoming_num.strip().lower()
    for row in range(ws.max_row, 1, -1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is not None and str(cell_val).strip().lower() == num_clean:
            hyperlink = ''
            try:
                hl = ws.cell(row=row, column=3).hyperlink
                if hl: hyperlink = hl.target if hasattr(hl, 'target') else str(hl)
            except Exception: pass
            def _str(v):
                if v is None: return ''
                if hasattr(v, 'strftime'): return v.strftime('%d.%m.%Y')
                return str(v).strip()
            return {
                'date': _str(ws.cell(row,1).value), 'incoming_num': _str(ws.cell(row,2).value),
                'letter_num': _str(ws.cell(row,3).value), 'subject': _str(ws.cell(row,4).value),
                'author': _str(ws.cell(row,5).value), 'signed_by': _str(ws.cell(row,6).value),
                'folder_num': _str(ws.cell(row,7).value), 'keywords': _str(ws.cell(row,9).value),
                'related': _str(ws.cell(row,10).value), 'hyperlink': hyperlink,
            }
    return None

# ── Запись в журнал ───────────────────────────────────────────────────────────

def write_to_excel_in(row_data: dict):
    if not HAS_OPENPYXL: raise RuntimeError("openpyxl не установлен")
    if not os.path.exists(EXCEL_PATH_IN):
        raise FileNotFoundError(f"Файл журнала не найден:\n{EXCEL_PATH_IN}")
    wb = openpyxl.load_workbook(EXCEL_PATH_IN)
    ws = wb.worksheets[-1]
    last_row = ws.max_row
    while last_row > 1 and ws.cell(row=last_row, column=1).value is None: last_row -= 1
    r = last_row + 1
    ws.cell(r,1).value = row_data['date']
    ws.cell(r,2).value = row_data['incoming_num']
    c3 = ws.cell(r,3); c3.value = row_data['letter_num']
    if row_data.get('hyperlink_path'): c3.hyperlink = row_data['hyperlink_path']; c3.style = 'Hyperlink'
    ws.cell(r,4).value = row_data['subject']
    ws.cell(r,5).value = row_data['author']
    c6 = ws.cell(r,6); c6.value = row_data['signed_by']; c6.alignment = Alignment(wrap_text=True)
    ws.cell(r,7).value = row_data['folder_num']
    ws.cell(r,8).value = row_data['who_registered']
    ws.cell(r,9).value = row_data['keywords']
    ws.cell(r,10).value = row_data['related']
    wb.save(EXCEL_PATH_IN)

def write_to_excel_out(row_data: dict):
    if not HAS_OPENPYXL: raise RuntimeError("openpyxl не установлен")
    if not os.path.exists(EXCEL_PATH_OUT):
        raise FileNotFoundError(f"Файл журнала не найден:\n{EXCEL_PATH_OUT}")
    wb = openpyxl.load_workbook(EXCEL_PATH_OUT)
    ws = wb.worksheets[-1]
    last_row = ws.max_row
    while last_row > 1 and ws.cell(row=last_row, column=1).value is None: last_row -= 1
    r = last_row + 1
    ws.cell(r,1).value = row_data['date']
    c2 = ws.cell(r,2); c2.value = row_data['letter_num']
    if row_data.get('hyperlink_path'): c2.hyperlink = row_data['hyperlink_path']; c2.style = 'Hyperlink'
    ws.cell(r,3).value = row_data['subject']
    c4 = ws.cell(r,4); c4.value = row_data['recipient']; c4.alignment = Alignment(wrap_text=True)
    ws.cell(r,5).value = row_data['executor']
    ws.cell(r,6).value = row_data['keywords']
    ws.cell(r,7).value = row_data['related']
    ws.cell(r,8).value = row_data['control']
    wb.save(EXCEL_PATH_OUT)

# ── Окно настроек ─────────────────────────────────────────────────────────────

class SettingsWindow(tk.Toplevel):

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Настройки")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()
        self._build()
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    def _build(self):
        f = ttk.Frame(self, padding=12)
        f.grid(row=0, column=0, sticky='nsew')
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        f.columnconfigure(1, weight=1)
        f.rowconfigure(4, weight=1)

        # Имя регистратора
        ttk.Label(f, text="Имя регистратора (столбец H входящих):").grid(
            row=0, column=0, sticky='e', padx=(0, 6), pady=(0, 4))
        self._reg_var = tk.StringVar(value=_settings.get('registrar_name', ''))
        ttk.Entry(f, textvariable=self._reg_var, width=32).grid(
            row=0, column=1, sticky='ew', pady=(0, 4))
        ttk.Label(f, text="(Оставьте пустым — будет использоваться системное имя пользователя)",
                  foreground='gray').grid(row=1, column=0, columnspan=2, sticky='w', pady=(0, 12))

        # Сокращения организаций
        ttk.Label(f, text="Сокращения организаций:").grid(
            row=2, column=0, columnspan=2, sticky='w', pady=(0, 2))
        ttk.Label(f, text="Формат:  Полное название организации = Сокращение",
                  foreground='gray').grid(row=3, column=0, columnspan=2, sticky='w', pady=(0, 4))

        txt_frame = ttk.Frame(f)
        txt_frame.grid(row=4, column=0, columnspan=2, sticky='nsew', pady=(0, 8))
        txt_frame.columnconfigure(0, weight=1)
        txt_frame.rowconfigure(0, weight=1)

        self._abbr_text = tk.Text(txt_frame, width=72, height=18, wrap='none',
                                   font=('Consolas', 9))
        sb_v = ttk.Scrollbar(txt_frame, orient='vertical', command=self._abbr_text.yview)
        sb_h = ttk.Scrollbar(txt_frame, orient='horizontal', command=self._abbr_text.xview)
        self._abbr_text.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
        self._abbr_text.grid(row=0, column=0, sticky='nsew')
        sb_v.grid(row=0, column=1, sticky='ns')
        sb_h.grid(row=1, column=0, sticky='ew')

        # Заполняем из словаря
        abbrs = _settings.get('org_abbreviations', {})
        for key, val in abbrs.items():
            self._abbr_text.insert('end', f"{key} = {val}\n")

        # Кнопки
        btn_frame = ttk.Frame(f)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(4, 0))
        ttk.Button(btn_frame, text="Сохранить", command=self._save).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side='left', padx=6)

    def _save(self):
        lines = self._abbr_text.get('1.0', 'end').strip().splitlines()
        abbrs = {}
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, _, val = line.partition('=')
                key, val = key.strip(), val.strip()
                if key and val:
                    abbrs[key] = val
        _settings['org_abbreviations'] = abbrs
        _settings['registrar_name'] = self._reg_var.get().strip()
        save_settings()
        self.destroy()

# ── Главное окно ──────────────────────────────────────────────────────────────

class RegistrationApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.in_data  = {}
        self.out_data = {}
        self._in_preview_vars  = {}
        self._out_preview_vars = {}
        self._default_filename_in  = ''
        self._default_filename_out = ''

        self.title("Регистрация корреспонденции")
        self.resizable(True, True)
        self._build_ui()
        self._center_window()

    # ── UI ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        root = ttk.Frame(self, padding=8)
        root.grid(row=0, column=0, sticky='nsew')
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)

        self._reg_nb = ttk.Notebook(root)
        self._reg_nb.grid(row=0, column=0, sticky='nsew')

        in_tab  = ttk.Frame(self._reg_nb, padding=8)
        out_tab = ttk.Frame(self._reg_nb, padding=8)
        self._reg_nb.add(in_tab,  text="  Входящие  ")
        self._reg_nb.add(out_tab, text="  Исходящие  ")
        self._build_incoming_tab(in_tab)
        self._build_outgoing_tab(out_tab)

        self._reparse_status = tk.StringVar(value="")
        ttk.Label(root, textvariable=self._reparse_status,
                  foreground='gray').grid(row=1, column=0, pady=(2, 0))
        btn_frame = ttk.Frame(root)
        btn_frame.grid(row=2, column=0, pady=4)
        ttk.Button(btn_frame, text="Зарегистрировать в журнал",
                   command=self._on_register).pack(side='left', padx=6)
        self._reparse_btn = ttk.Button(btn_frame, text="Считать из LanDocs заново",
                                       command=self._start_reparse)
        self._reparse_btn.pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Настройки",
                   command=self._open_settings).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Закрыть",
                   command=self.destroy).pack(side='left', padx=6)

    def _build_incoming_tab(self, frame):
        frame.columnconfigure(1, weight=1)

        # Импорт из журнала
        imp = ttk.LabelFrame(frame, text="Импорт из журнала (по № вх)", padding=8)
        imp.grid(row=0, column=0, sticky='ew', pady=(0, 6))
        imp.columnconfigure(3, weight=1)
        ttk.Label(imp, text="№ вх:").grid(row=0, column=0, sticky='e', padx=(0,4))
        self._in_import_num = tk.StringVar()
        ttk.Entry(imp, textvariable=self._in_import_num, width=18).grid(row=0, column=1)
        ttk.Label(imp, text="  Год:").grid(row=0, column=2, sticky='e', padx=(8,4))
        self._in_import_year = tk.StringVar(value=str(datetime.now().year))
        ttk.Entry(imp, textvariable=self._in_import_year, width=6).grid(row=0, column=3, sticky='w')
        ttk.Button(imp, text="Найти в журнале",
                   command=self._on_import_journal).grid(row=0, column=4, padx=(8,0))
        self._import_status = tk.StringVar(value="")
        ttk.Label(imp, textvariable=self._import_status,
                  foreground='navy').grid(row=1, column=0, columnspan=5, pady=(4,0), sticky='w')

        # Данные из LanDocs
        info = ttk.LabelFrame(frame, text="Данные из LanDocs", padding=8)
        info.grid(row=1, column=0, sticky='ew', pady=(0, 8))
        info.columnconfigure(1, weight=1)
        for i, (label, key) in enumerate([
            ("Дата:","date"),("№ вх:","incoming_num"),("№ письма:","letter_num"),
            ("Тема письма:","subject"),("Подписант:","signatory"),
            ("Корреспондент:","correspondent"),("Связанное письмо:","related"),
        ]):
            ttk.Label(info, text=label, anchor='e').grid(row=i, column=0, sticky='e', padx=(0,6), pady=2)
            var = tk.StringVar(); self._in_preview_vars[key] = var
            ttk.Label(info, textvariable=var, anchor='w', wraplength=420).grid(row=i, column=1, sticky='w', pady=2)

        # Поля для заполнения
        inp = ttk.LabelFrame(frame, text="Данные для регистрации", padding=8)
        inp.grid(row=2, column=0, sticky='ew')
        inp.columnconfigure(1, weight=1)
        ttk.Label(inp, text="Автор письма:").grid(row=0, column=0, sticky='e', padx=(0,6), pady=4)
        self.in_author_var = tk.StringVar(value="-")
        ttk.Entry(inp, textvariable=self.in_author_var, width=48).grid(row=0, column=1, sticky='ew', pady=4)
        ttk.Label(inp, text="Ключевые слова:").grid(row=1, column=0, sticky='e', padx=(0,6), pady=4)
        self.in_keywords_var = tk.StringVar()
        ttk.Entry(inp, textvariable=self.in_keywords_var, width=48).grid(row=1, column=1, sticky='ew', pady=4)
        ttk.Label(inp, text="Название файла:").grid(row=2, column=0, sticky='e', padx=(0,6), pady=4)
        self.in_filename_var = tk.StringVar()
        ttk.Entry(inp, textvariable=self.in_filename_var, width=48).grid(row=2, column=1, sticky='ew', pady=4)
        ttk.Label(inp, text="Папка сохранения:").grid(row=3, column=0, sticky='e', padx=(0,6), pady=4)
        row3 = ttk.Frame(inp); row3.grid(row=3, column=1, sticky='ew', pady=4)
        row3.columnconfigure(0, weight=1)
        self.in_save_path_var = tk.StringVar()
        ttk.Entry(row3, textvariable=self.in_save_path_var, state='readonly', width=38).grid(row=0, column=0, sticky='ew')
        ttk.Button(row3, text="Выбрать папку…",
                   command=self._choose_save_folder_in).grid(row=0, column=1, padx=(6,0))
        ttk.Label(inp, text="№ папки:").grid(row=4, column=0, sticky='e', padx=(0,6), pady=4)
        self.in_folder_num_var = tk.StringVar()
        ttk.Label(inp, textvariable=self.in_folder_num_var,
                  anchor='w', foreground='navy').grid(row=4, column=1, sticky='w', pady=4)

    def _build_outgoing_tab(self, frame):
        frame.columnconfigure(1, weight=1)
        info = ttk.LabelFrame(frame, text="Данные из LanDocs", padding=8)
        info.grid(row=0, column=0, sticky='ew', pady=(0, 8))
        info.columnconfigure(1, weight=1)
        for i, (label, key) in enumerate([
            ("Дата:","date"),("№ письма:","letter_num"),("Тема письма:","subject"),
            ("Исполнитель:","executor"),("Получатели (ФИО):","recipient_names"),
            ("Получатели (орг):","recipient_companies"),("Связанное письмо:","related"),
        ]):
            ttk.Label(info, text=label, anchor='e').grid(row=i, column=0, sticky='e', padx=(0,6), pady=2)
            var = tk.StringVar(); self._out_preview_vars[key] = var
            ttk.Label(info, textvariable=var, anchor='w', wraplength=420).grid(row=i, column=1, sticky='w', pady=2)
        inp = ttk.LabelFrame(frame, text="Данные для регистрации", padding=8)
        inp.grid(row=1, column=0, sticky='ew')
        inp.columnconfigure(1, weight=1)
        ttk.Label(inp, text="Ключевые слова:").grid(row=0, column=0, sticky='e', padx=(0,6), pady=4)
        self.out_keywords_var = tk.StringVar()
        ttk.Entry(inp, textvariable=self.out_keywords_var, width=48).grid(row=0, column=1, sticky='ew', pady=4)
        ttk.Label(inp, text="Контроль:").grid(row=1, column=0, sticky='e', padx=(0,6), pady=4)
        self.out_control_var = tk.StringVar()
        ttk.Entry(inp, textvariable=self.out_control_var, width=48).grid(row=1, column=1, sticky='ew', pady=4)
        ttk.Label(inp, text="Название файла:").grid(row=2, column=0, sticky='e', padx=(0,6), pady=4)
        self.out_filename_var = tk.StringVar()
        ttk.Entry(inp, textvariable=self.out_filename_var, width=48).grid(row=2, column=1, sticky='ew', pady=4)
        ttk.Label(inp, text="Папка сохранения:").grid(row=3, column=0, sticky='e', padx=(0,6), pady=4)
        row3 = ttk.Frame(inp); row3.grid(row=3, column=1, sticky='ew', pady=4)
        row3.columnconfigure(0, weight=1)
        self.out_save_path_var = tk.StringVar()
        ttk.Entry(row3, textvariable=self.out_save_path_var, state='readonly', width=38).grid(row=0, column=0, sticky='ew')
        ttk.Button(row3, text="Выбрать папку…",
                   command=self._choose_save_folder_out).grid(row=0, column=1, padx=(6,0))
        ttk.Label(inp, text="№ папки:").grid(row=4, column=0, sticky='e', padx=(0,6), pady=4)
        self.out_folder_num_var = tk.StringVar()
        ttk.Label(inp, textvariable=self.out_folder_num_var,
                  anchor='w', foreground='navy').grid(row=4, column=1, sticky='w', pady=4)

    # ── Применение данных ─────────────────────────────────────────────────

    def _apply_incoming_data(self):
        d = self.in_data
        for key, var in self._in_preview_vars.items():
            var.set(d.get(key, ''))
        correspondent = d.get('correspondent', '')
        self._default_filename_in = build_default_filename_in(
            d.get('date',''), d.get('incoming_num',''), d.get('letter_num',''),
            d.get('subject',''), correspondent)
        self.in_filename_var.set(self._default_filename_in)
        self.in_keywords_var.set(d.get('subject', ''))
        self.in_save_path_var.set('')
        self.in_folder_num_var.set('')
        # Автор: если есть сокращение — «Сокращение-», иначе «-»
        abbr = abbreviate_org(correspondent) if correspondent else ''
        if abbr and abbr != correspondent.strip():
            self.in_author_var.set(f"{abbr}-")
        else:
            self.in_author_var.set('-')

    def _apply_outgoing_data(self):
        d = self.out_data
        for key, var in self._out_preview_vars.items():
            var.set(d.get(key, ''))
        self._default_filename_out = build_default_filename_out(
            d.get('date',''), d.get('letter_num',''),
            d.get('recipient_companies',''), d.get('subject',''))
        self.out_filename_var.set(self._default_filename_out)
        self.out_keywords_var.set(d.get('subject',''))
        self.out_save_path_var.set('')
        self.out_folder_num_var.set('')

    # ── Импорт из журнала ─────────────────────────────────────────────────

    def _on_import_journal(self):
        num  = self._in_import_num.get().strip()
        year = self._in_import_year.get().strip()
        if not num:
            self._import_status.set("Введите № вх"); return
        try:
            result = lookup_incoming_journal(num, year)
        except Exception as exc:
            self._import_status.set(f"Ошибка: {exc}"); return
        if result is None:
            self._import_status.set(f"Не найдено: {num} за {year} год"); return
        self.in_data = {
            'incoming_num': result.get('incoming_num', num),
            'date':         result.get('date', ''),
            'letter_num':   result.get('letter_num', ''),
            'subject':      result.get('subject', ''),
            'signatory':    '',
            'correspondent': result.get('signed_by', ''),
            'related':      result.get('related', ''),
            'file_link':    result.get('hyperlink', ''),
        }
        self._apply_incoming_data()
        self.in_author_var.set(result.get('author', '-'))
        self.in_keywords_var.set(result.get('keywords', ''))
        self._import_status.set(f"Импортировано: вх-{num}")

    # ── Парсинг LanDocs ───────────────────────────────────────────────────

    def _active_tab(self) -> str:
        return 'out' if self._reg_nb.index(self._reg_nb.select()) == 1 else 'in'

    def _start_reparse(self):
        self._reparse_btn.config(state='disabled')
        self.iconify()
        self._reparse_countdown(3)

    def _reparse_countdown(self, n: int):
        if n > 0:
            self._reparse_status.set(f"Переключитесь в LanDocs… считывание через {n} сек.")
            self.after(1000, self._reparse_countdown, n - 1)
        else:
            self._reparse_status.set("Выполняется считывание…")
            self.after(50, self._do_reparse)

    def _do_reparse(self):
        tab = self._active_tab()
        try:
            if tab == 'in':
                self.in_data = extract_landocs_data_in()
                self._apply_incoming_data()
            else:
                self.out_data = extract_landocs_data_out()
                self._apply_outgoing_data()
            self._reparse_status.set("Считывание завершено.")
        except Exception as exc:
            self._reparse_status.set(f"Ошибка: {exc}")
        finally:
            self._reparse_btn.config(state='normal')
            self.deiconify(); self.lift()

    # ── Выбор файла сохранения ────────────────────────────────────────────

    def _choose_save_folder_in(self):
        self._choose_save_folder(
            file_link=self.in_data.get('file_link', ''),
            filename_var=self.in_filename_var,
            default_filename=self._default_filename_in,
            save_path_var=self.in_save_path_var,
            folder_num_var=self.in_folder_num_var,
        )

    def _choose_save_folder_out(self):
        self._choose_save_folder(
            file_link=self.out_data.get('file_link', ''),
            filename_var=self.out_filename_var,
            default_filename=self._default_filename_out,
            save_path_var=self.out_save_path_var,
            folder_num_var=self.out_folder_num_var,
        )

    def _choose_save_folder(self, file_link, filename_var, default_filename,
                            save_path_var, folder_num_var):
        initial = DEFAULT_SAVE_FOLDER if os.path.isdir(DEFAULT_SAVE_FOLDER) else os.path.expanduser("~")
        # Определяем расширение: из ссылки LanDocs → ViewDir → .pdf
        ext = os.path.splitext(file_link)[1].lower() if file_link else ''
        if not ext:
            latest = find_latest_in_viewdir()
            ext = os.path.splitext(latest)[1].lower() if latest else ''
        if not ext or ext not in LETTER_EXTS:
            ext = '.pdf'
        filename_base = filename_var.get() or default_filename
        selected = filedialog.asksaveasfilename(
            title="Выберите папку и имя файла для сохранения письма",
            initialdir=initial,
            initialfile=filename_base + ext,
            defaultextension=ext,
            filetypes=LETTER_FILETYPES,
        )
        if not selected: return
        selected = selected.replace('/', '\\')
        save_path_var.set(selected)
        base_name = os.path.splitext(os.path.basename(selected))[0]
        if base_name: filename_var.set(base_name)
        folder_num_var.set(calc_folder_num(os.path.dirname(selected)))

    # ── Регистрация ───────────────────────────────────────────────────────

    def _on_register(self):
        if self._active_tab() == 'in':
            self._on_register_in()
        else:
            self._on_register_out()

    def _on_register_in(self):
        if not self.in_save_path_var.get():
            messagebox.showwarning("Внимание", "Выберите папку и имя файла.", parent=self); return
        if not HAS_OPENPYXL:
            messagebox.showerror("Ошибка", "openpyxl не установлен.\npip install openpyxl", parent=self); return
        try:
            d = self.in_data
            signatory     = abbreviate_fio(d.get('signatory', ''))
            correspondent = abbreviate_org(d.get('correspondent', ''))
            signed_by = f"{signatory}\n{correspondent}" if correspondent else signatory
            hyperlink_path = self.in_save_path_var.get()
            if hyperlink_path:
                src = find_latest_in_viewdir()
                if src:
                    shutil.copy2(src, hyperlink_path)
                else:
                    raise FileNotFoundError(
                        "Файл письма не найден в ViewDir.\n"
                        r"Убедитесь, что письмо открыто в LanDocs (%LOCALAPPDATA%\Temp\ViewDir)")
            registrar = _settings.get('registrar_name', '').strip() or getpass.getuser()
            write_to_excel_in({
                'date':           fmt_date_ymd(d.get('date','')),
                'incoming_num':   d.get('incoming_num',''),
                'letter_num':     d.get('letter_num',''),
                'subject':        d.get('subject',''),
                'author':         self.in_author_var.get(),
                'signed_by':      signed_by,
                'folder_num':     self.in_folder_num_var.get(),
                'who_registered': registrar,
                'keywords':       self.in_keywords_var.get(),
                'related':        d.get('related',''),
                'hyperlink_path': hyperlink_path,
            })
            messagebox.showinfo("Готово", "Запись добавлена в журнал!", parent=self)
            # Сброс пути после регистрации (готов к следующему письму)
            self.in_save_path_var.set('')
            self.in_folder_num_var.set('')
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc), parent=self)

    def _on_register_out(self):
        if not self.out_save_path_var.get():
            messagebox.showwarning("Внимание", "Выберите папку и имя файла.", parent=self); return
        if not HAS_OPENPYXL:
            messagebox.showerror("Ошибка", "openpyxl не установлен.\npip install openpyxl", parent=self); return
        try:
            d = self.out_data
            hyperlink_path = self.out_save_path_var.get()
            if hyperlink_path:
                src = find_latest_in_viewdir()
                if src:
                    shutil.copy2(src, hyperlink_path)
                else:
                    raise FileNotFoundError("Файл письма не найден в ViewDir.")
            write_to_excel_out({
                'date':           fmt_date_ymd(d.get('date','')),
                'letter_num':     d.get('letter_num',''),
                'subject':        d.get('subject',''),
                'recipient':      build_recipient_string(
                                      d.get('recipient_names',''), d.get('recipient_companies','')),
                'executor':       abbreviate_fio(d.get('executor','')),
                'keywords':       self.out_keywords_var.get(),
                'related':        d.get('related',''),
                'control':        self.out_control_var.get(),
                'hyperlink_path': hyperlink_path,
            })
            messagebox.showinfo("Готово", "Запись добавлена в журнал!", parent=self)
            self.out_save_path_var.set('')
            self.out_folder_num_var.set('')
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc), parent=self)

    # ── Настройки ─────────────────────────────────────────────────────────

    def _open_settings(self):
        SettingsWindow(self)

    # ── Центровка ─────────────────────────────────────────────────────────

    def _center_window(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")


# ── Точка входа ───────────────────────────────────────────────────────────────

def main():
    load_settings()
    missing = []
    if not HAS_WIN32:    missing.append("pywin32")
    if not HAS_OPENPYXL: missing.append("openpyxl")
    if missing:
        root = tk.Tk(); root.withdraw()
        messagebox.showerror("Не хватает зависимостей",
            "Не установлены:\n  " + "\n  ".join(missing) +
            "\n\nПрограмма запустится в демо-режиме (без LanDocs).")
        root.destroy()
        if not HAS_WIN32:
            app = RegistrationApp()
            app.mainloop()
            return
        sys.exit(1)
    time.sleep(0.4)
    app = RegistrationApp()
    app.mainloop()


if __name__ == '__main__':
    main()
